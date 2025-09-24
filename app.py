import streamlit as st
import pandas as pd
import io
from openpyxl.styles import PatternFill
from openpyxl import Workbook
import tempfile
import os

# Configuração da página
st.set_page_config(
    page_title="Processador de Relatórios Escolares",
    page_icon="📊",
    layout="wide"
)

# CSS personalizado
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1E3A8A;
        text-align: center;
        margin-bottom: 2rem;
    }
    .success-box {
        background-color: #D4EDDA;
        color: #155724;
        padding: 15px;
        border-radius: 5px;
        margin: 10px 0;
    }
    .info-box {
        background-color: #D1ECF1;
        color: #0C5460;
        padding: 15px;
        border-radius: 5px;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)

def processar_compilacao(uploaded_files):
    """Função para compilar múltiplas planilhas"""
    try:
        dfs = []
        
        for uploaded_file in uploaded_files:
            # Processar cada planilha
            df = pd.read_excel(uploaded_file, header=None)
            df = df.drop(index=0).reset_index(drop=True)
            df.columns = df.iloc[0]
            df = df.drop(index=0).reset_index(drop=True)
            dfs.append(df)
        
        # Compilar todas as planilhas
        df_compilado = pd.concat(dfs, ignore_index=True)
        df_compilado = df_compilado.loc[:,~df_compilado.columns.duplicated()]
        
        return df_compilado, None
        
    except Exception as e:
        return None, f"Erro na compilação: {str(e)}"

def processar_relatorio_colorido(uploaded_file):
    """Função para processar relatório com cores"""
    try:
        # Ler o arquivo
        df = pd.read_excel(uploaded_file, sheet_name=0)
        
        # Verificar estrutura e encontrar cabeçalhos
        primeira_linha = df.iloc[0].tolist() if len(df) > 0 else []
        
        if 'DR' in df.columns or 'Nome' in df.columns:
            st.success("✅ Cabeçalhos identificados corretamente")
        else:
            # Buscar linha com cabeçalhos corretos
            header_found = False
            for i in range(min(5, len(df))):
                linha = df.iloc[i].tolist()
                if any('DR' in str(item) for item in linha) or any('Nome' in str(item) for item in linha):
                    df = pd.read_excel(uploaded_file, sheet_name=0, header=i)
                    st.success(f"✅ Cabeçalhos encontrados na linha {i+1}")
                    header_found = True
                    break
            
            if not header_found:
                # Mapeamento manual
                if len(df.columns) >= 11:
                    mapeamento_colunas = {
                        df.columns[0]: 'DR',
                        df.columns[1]: 'Polo', 
                        df.columns[2]: 'Nome',
                        df.columns[3]: 'Etapa',
                        df.columns[4]: 'Sala',
                        df.columns[5]: 'Área de conhecimento',
                        df.columns[6]: 'Atividades(tentativas/quantidade de tentativas)',
                        df.columns[7]: 'Menção Atual',
                        df.columns[8]: 'Data último acesso',
                        df.columns[9]: 'Brasileiro(a)',
                        df.columns[10]: 'Aluno AEE'
                    }
                    df = df.rename(columns=mapeamento_colunas)
                    st.info("⚠️ Usando mapeamento manual de colunas")
        
        # Limpar dados
        df = df.dropna(how='all')
        
        # Verificar colunas necessárias
        colunas_necessarias = ['Nome', 'Atividades(tentativas/quantidade de tentativas)', 'Menção Atual']
        colunas_faltantes = [col for col in colunas_necessarias if col not in df.columns]
        
        if colunas_faltantes:
            return None, f"Colunas faltantes: {', '.join(colunas_faltantes)}"
        
        # Criar identificador do aluno
        if 'Polo' in df.columns and 'Nome' in df.columns:
            df['Aluno_ID'] = df['Polo'] + ' - ' + df['Nome']
        elif 'Nome' in df.columns:
            df['Aluno_ID'] = df['Nome']
        else:
            df['Aluno_ID'] = 'Aluno_' + (df.index + 1).astype(str)
        
        # Processar atividades
        df['Atividade'] = df['Atividades(tentativas/quantidade de tentativas)'].astype(str).str.split('(').str[0].str.strip()
        
        # Extrair tentativas
        tentativas = df['Atividades(tentativas/quantidade de tentativas)'].astype(str).str.extract(r'\((\d+)/(\d+)\)')
        if not tentativas.empty:
            df['Tentativas_Realizadas'] = tentativas[0].fillna(0).astype(int)
            df['Tentativas_Total'] = tentativas[1].fillna(0).astype(int)
        
        # Pivot das menções
        pivot_mencoes = df.pivot_table(
            index='Aluno_ID',
            columns='Atividade',
            values='Menção Atual',
            aggfunc='first',
            fill_value='--'
        ).reset_index()
        
        # Informações do aluno
        colunas_aluno = ['Aluno_ID', 'DR', 'Polo', 'Nome', 'Etapa', 'Sala', 'Área de conhecimento', 
                        'Data último acesso', 'Brasileiro(a)', 'Aluno AEE']
        colunas_aluno = [col for col in colunas_aluno if col in df.columns]
        info_alunos = df[colunas_aluno].drop_duplicates(subset=['Aluno_ID'])
        
        # Combinar dados
        resultado = info_alunos.merge(pivot_mencoes, on='Aluno_ID', how='left')
        
        # Pivot das tentativas (se disponível)
        if 'Tentativas_Realizadas' in df.columns:
            pivot_tentativas = df.pivot_table(
                index='Aluno_ID',
                columns='Atividade',
                values='Tentativas_Realizadas',
                aggfunc='first',
                fill_value=0
            ).reset_index()
            pivot_tentativas.columns = ['Aluno_ID'] + [f'{col}_Tentativas' for col in pivot_tentativas.columns if col != 'Aluno_ID']
            resultado = resultado.merge(pivot_tentativas, on='Aluno_ID', how='left')
        
        # Reordenar colunas
        colunas_ordenadas = ['Aluno_ID', 'DR', 'Polo', 'Nome', 'Etapa', 'Sala', 'Área de conhecimento', 
                            'Data último acesso', 'Brasileiro(a)', 'Aluno AEE']
        colunas_ordenadas = [col for col in colunas_ordenadas if col in resultado.columns]
        colunas_atividades = [col for col in resultado.columns if col not in colunas_ordenadas and col != 'Aluno_ID']
        colunas_ordenadas.extend(colunas_atividades)
        resultado = resultado[colunas_ordenadas]
        
        return resultado, None
        
    except Exception as e:
        return None, f"Erro no processamento: {str(e)}"

def aplicar_cores_excel(df, nome_arquivo):
    """Aplicar cores alternadas no Excel"""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Relatorio_Processado', index=False)
            
            workbook = writer.book
            sheet = workbook['Relatorio_Processado']
            
            cor1 = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            cor2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            aluno_ids = df['Aluno_ID'].tolist()
            ultima_linha = sheet.max_row
            ultima_coluna = sheet.max_column
            
            cor_atual = cor1
            aluno_atual = None
            
            for i, linha in enumerate(range(2, ultima_linha + 1)):
                if i < len(aluno_ids) and aluno_ids[i] != aluno_atual:
                    cor_atual = cor2 if cor_atual == cor1 else cor1
                    aluno_atual = aluno_ids[i]
                
                for col in range(1, ultima_coluna + 1):
                    sheet.cell(row=linha, column=col).fill = cor_atual
        
        return tmp.name

# Interface principal
st.markdown('<div class="main-header">📊 PROCESSADOR DE RELATÓRIOS ESCOLARES</div>', unsafe_allow_html=True)

# Sidebar com informações
with st.sidebar:
    st.header("ℹ️ Informações")
    st.info("""
    **Funcionalidades:**
    - 📥 **Compilar Planilhas**: Une múltiplos relatórios em um único arquivo
    - 🎨 **Relatório Colorido**: Processa e formata com cores alternadas por aluno
    - 📊 **Visualização**: Preview dos dados processados
    """)
    
    st.header("📋 Instruções")
    st.write("""
    1. Selecione a funcionalidade desejada
    2. Faça upload do(s) arquivo(s) Excel
    3. Aguarde o processamento
    4. Baixe o resultado
    """)

# Seleção de funcionalidade
funcionalidade = st.radio(
    "Selecione a funcionalidade:",
    ["Compilar Múltiplas Planilhas", "Processar Relatório Colorido"],
    horizontal=True
)

if funcionalidade == "Compilar Múltiplas Planilhas":
    st.subheader("📥 Compilar Múltiplas Planilhas")
    
    uploaded_files = st.file_uploader(
        "Selecione as planilhas para compilar:",
        type=['xlsx'],
        accept_multiple_files=True,
        help="Selecione todos os arquivos Excel que deseja compilar"
    )
    
    if uploaded_files:
        if st.button("🚀 Compilar Planilhas"):
            with st.spinner("Processando compilação..."):
                df_compilado, erro = processar_compilacao(uploaded_files)
                
                if erro:
                    st.error(erro)
                else:
                    st.success("✅ Compilação concluída com sucesso!")
                    
                    # Estatísticas
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total de Linhas", len(df_compilado))
                    with col2:
                        st.metric("Total de Colunas", len(df_compilado.columns))
                    with col3:
                        st.metric("Arquivos Compilados", len(uploaded_files))
                    
                    # Preview
                    st.subheader("👀 Preview dos Dados Compilados")
                    st.dataframe(df_compilado.head(10))
                    
                    # Download
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_compilado.to_excel(writer, index=False)
                    
                    st.download_button(
                        label="📥 Download do Arquivo Compilado",
                        data=output.getvalue(),
                        file_name="relatorio_compilado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

else:  # Processar Relatório Colorido
    st.subheader("🎨 Processar Relatório Colorido")
    
    uploaded_file = st.file_uploader(
        "Selecione o arquivo Excel para processar:",
        type=['xlsx'],
        help="Selecione um arquivo Excel com o relatório escolar"
    )
    
    if uploaded_file:
        if st.button("🚀 Processar Relatório"):
            with st.spinner("Processando relatório..."):
                resultado, erro = processar_relatorio_colorido(uploaded_file)
                
                if erro:
                    st.error(erro)
                else:
                    st.success("✅ Relatório processado com sucesso!")
                    
                    # Estatísticas
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total de Alunos", resultado['Aluno_ID'].nunique())
                    with col2:
                        st.metric("Total de Atividades", len([col for col in resultado.columns if col not in ['Aluno_ID', 'DR', 'Polo', 'Nome', 'Etapa', 'Sala', 'Área de conhecimento', 'Data último acesso', 'Brasileiro(a)', 'Aluno AEE']]))
                    with col3:
                        st.metric("Total de Colunas", len(resultado.columns))
                    
                    # Preview
                    st.subheader("👀 Preview do Relatório Processado")
                    st.dataframe(resultado.head(10))
                    
                    # Aplicar cores e criar download
                    with st.spinner("Aplicando formatação de cores..."):
                        arquivo_temp = aplicar_cores_excel(resultado, uploaded_file.name)
                        
                        with open(arquivo_temp, 'rb') as f:
                            arquivo_bytes = f.read()
                        
                        st.download_button(
                            label="📥 Download do Relatório Colorido",
                            data=arquivo_bytes,
                            file_name=uploaded_file.name.replace('.xlsx', '_PROCESSADO_COLORIDO.xlsx'),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    # Limpar arquivo temporário
                    os.unlink(arquivo_temp)

# Rodapé
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: #666;'>"
    "Desenvolvido para processamento de relatórios escolares | "
    "Versão 1.0"
    "</div>", 
    unsafe_allow_html=True
)
